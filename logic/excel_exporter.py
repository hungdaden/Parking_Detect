import io
from typing import Any, Dict
from PIL import Image as PILImage

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime


def export_parking_report_to_excel(file_path: str, db_manager: Any, filter_opts: Dict[str, Any]) -> str:
    """
    Exports parking database reports to an Excel file with ClickUp-inspired clean styling and embedded plate images.

    filter_opts dict structure:
    {
        "start_dt": "YYYY-MM-DD HH:MM:SS" or None,
        "end_dt": "YYYY-MM-DD HH:MM:SS" or None,
        "start_time": "HH:MM:SS" or None,
        "end_time": "HH:MM:SS" or None,
        "filter_description": "Custom text describing the filter range"
    }
    """
    start_dt = filter_opts.get("start_dt")
    end_dt = filter_opts.get("end_dt")
    start_time = filter_opts.get("start_time")
    end_time = filter_opts.get("end_time")
    filter_desc = filter_opts.get("filter_description", "Tất cả dữ liệu")

    # Retrieve filtered data from DB
    events = db_manager.get_filtered_events(start_dt, end_dt, start_time, end_time)
    misparked_events = db_manager.get_filtered_misparked_events(start_dt, end_dt, start_time, end_time)
    slot_summary = db_manager.get_filtered_slot_summary(start_dt, end_dt, start_time, end_time)
    license_plates = db_manager.get_filtered_license_plates(start_dt, end_dt, start_time, end_time)
    stats = db_manager.get_filtered_stats(start_dt, end_dt, start_time, end_time)

    wb = openpyxl.Workbook()

    # ClickUp-inspired Color Palette
    font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, color="1E293B")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    
    # Status badges fonts & fills
    font_badge_in = Font(name="Segoe UI", size=10, bold=True, color="15803D")
    fill_badge_in = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    font_badge_out = Font(name="Segoe UI", size=10, bold=True, color="B45309")
    fill_badge_out = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    font_badge_misparked = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
    fill_badge_misparked = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    fill_title = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # ClickUp Indigo
    fill_header = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid") # Soft Indigo Header
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_kpi = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="E2E8F0")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # ==========================================
    # SHEET 1: TỔNG QUAN (OVERVIEW)
    # ==========================================
    ws_summary: Worksheet = wb.active  # type: ignore
    ws_summary.title = "Tổng Quan & Thống Kê"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    title_cell = ws_summary.cell(row=1, column=1, value="⚡ BÁO CÁO THỐNG KÊ HOẠT ĐỘNG BÃI ĐỖ XE")
    ws_summary.merge_cells("A1:E2")
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = align_center

    # Metadata Info
    cell_a4 = ws_summary.cell(row=4, column=1, value="Thời gian xuất báo cáo:")
    cell_a4.font = font_bold
    cell_b4 = ws_summary.cell(row=4, column=2, value=now_str)
    cell_b4.font = font_data

    cell_a5 = ws_summary.cell(row=5, column=1, value="Phạm vi lọc dữ liệu:")
    cell_a5.font = font_bold
    cell_b5 = ws_summary.cell(row=5, column=2, value=filter_desc)
    cell_b5.font = font_subtitle

    # KPI Summary Cards
    cell_a7 = ws_summary.cell(row=7, column=1, value="📊 THỐNG KÊ TỔNG QUAN")
    cell_a7.font = font_section

    kpi_headers = ["Tổng Lượt Xe Vào", "Tổng Lượt Xe Ra", "Tổng Số Sự Kiện", "Đỗ Sai Vị Trí"]
    kpi_values = [stats["total_in"], stats["total_out"], len(events), stats.get("total_misparked", 0)]

    for idx, (h, v) in enumerate(zip(kpi_headers, kpi_values), start=1):
        cell_h = ws_summary.cell(row=8, column=idx, value=h)
        cell_h.font = font_bold
        cell_h.fill = fill_kpi
        cell_h.alignment = align_center

        cell_v = ws_summary.cell(row=9, column=idx, value=v)
        cell_v.font = Font(name="Segoe UI", size=14, bold=True, color="E11D48" if idx == 4 else "4F46E5")
        cell_v.fill = fill_kpi
        cell_v.alignment = align_center

    # Slot Breakdown Table
    cell_a12 = ws_summary.cell(row=12, column=1, value="🅿️ CHI TIẾT LƯỢT XE THEO TỪNG Ô ĐỖ")
    cell_a12.font = font_section

    headers_slot = ["Mã Ô Đỗ", "Lượt Xe Vào", "Lượt Xe Ra"]
    for col_idx, h in enumerate(headers_slot, start=1):
        cell = ws_summary.cell(row=13, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    start_row = 14
    for i, s in enumerate(slot_summary):
        r = start_row + i
        c1 = ws_summary.cell(row=r, column=1, value=f"Ô đỗ #{s['slot_id']}")
        c2 = ws_summary.cell(row=r, column=2, value=s['total_in'])
        c3 = ws_summary.cell(row=r, column=3, value=s['total_out'])

        for cell in [c1, c2, c3]:
            cell.font = font_data
            cell.alignment = align_center
            cell.border = border_all
            if i % 2 == 1:
                cell.fill = fill_zebra

    if not slot_summary:
        empty_c = ws_summary.cell(row=14, column=1, value="Chưa có dữ liệu ô đỗ nào.")
        empty_c.font = font_subtitle

    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column if col[0].column is not None else 1)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ==========================================
    # SHEET 2: LỊCH SỬ XE VÀO RA (EVENTS LOG)
    # ==========================================
    ws_events: Worksheet = wb.create_sheet(title="Lịch Sử Xe Vào Ra")
    ws_events.views.sheetView[0].showGridLines = True

    # Title Banner
    t_ev = ws_events.cell(row=1, column=1, value="🚗 DANH SÁCH CHI TIẾT LỊCH SỬ XE VÀO / RA BÃI")
    ws_events.merge_cells("A1:F2")
    t_ev.font = font_title
    t_ev.fill = fill_title
    t_ev.alignment = align_center

    cell_ev_a4 = ws_events.cell(row=4, column=1, value=f"Bộ lọc: {filter_desc}")
    cell_ev_a4.font = font_subtitle

    event_headers = ["STT", "ID Sự Kiện", "Thời Gian", "Vị Trí Ô", "Mã Định Danh Xe", "Loại Sự Kiện"]
    for col_idx, text in enumerate(event_headers, start=1):
        cell = ws_events.cell(row=6, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    for idx, ev in enumerate(events, start=1):
        r = 6 + idx
        is_in = (ev["event_type"] == "IN")
        event_type_str = "XE VÀO (IN)" if is_in else "XE RA (OUT)"
        vals = [idx, ev["id"], ev["timestamp"], f"Ô {ev['slot_id']}", ev["vehicle_id"] or "", event_type_str]

        for col_idx, val in enumerate(vals, start=1):
            cell = ws_events.cell(row=r, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = align_center
            cell.border = border_all
            if r % 2 == 0 and col_idx != 6:
                cell.fill = fill_zebra
            if col_idx == 6:
                cell.font = font_badge_in if is_in else font_badge_out
                cell.fill = fill_badge_in if is_in else fill_badge_out

        ws_events.row_dimensions[r].height = 20

    if not events:
        empty_ev = ws_events.cell(row=7, column=1, value="Không tìm thấy sự kiện nào.")
        empty_ev.font = font_subtitle

    for col in ws_events.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column if col[0].column is not None else 1)
        ws_events.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ==========================================
    # SHEET 3: LỊCH SỬ BIỂN SỐ (ALPR LOGS WITH IMAGES)
    # ==========================================
    ws_plates: Worksheet = wb.create_sheet(title="Lịch Sử Biển Số ALPR")
    ws_plates.views.sheetView[0].showGridLines = True

    # Title Banner
    t_pl = ws_plates.cell(row=1, column=1, value="🔍 NGÀY GIỜ VÀ KẾT QUẢ NHẬN DIỆN BIỂN SỐ XE (ALPR)")
    ws_plates.merge_cells("A1:D2")
    t_pl.font = font_title
    t_pl.fill = fill_title
    t_pl.alignment = align_center

    cell_pl_a4 = ws_plates.cell(row=4, column=1, value=f"Bộ lọc: {filter_desc}")
    cell_pl_a4.font = font_subtitle

    plate_headers = ["STT", "Thời Gian Ghi Nhận", "Biển Số Nhận Diện", "Hình Ảnh Chụp Biển Số"]
    for col_idx, text in enumerate(plate_headers, start=1):
        cell = ws_plates.cell(row=6, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    for idx, p in enumerate(license_plates, start=1):
        r = 6 + idx
        vals = [idx, p["timestamp"], p["plate_text"]]

        for col_idx, val in enumerate(vals, start=1):
            cell = ws_plates.cell(row=r, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = align_center
            cell.border = border_all
            if col_idx == 3:
                cell.font = Font(name="Segoe UI", size=11, bold=True, color="4F46E5")
            if r % 2 == 0:
                cell.fill = fill_zebra

        # Column 4: License Plate Image Embedding
        cell_img: Any = ws_plates.cell(row=r, column=4)
        cell_img.border = border_all
        if r % 2 == 0:
            cell_img.fill = fill_zebra

        plate_img_bytes = p.get("plate_image")
        if plate_img_bytes:
            try:
                pil_img = PILImage.open(io.BytesIO(plate_img_bytes))
                orig_w, orig_h = pil_img.size
                target_h = 40
                target_w = int(orig_w * (target_h / orig_h)) if orig_h > 0 else 100

                img_stream = io.BytesIO()
                pil_img.save(img_stream, format="PNG")
                img_stream.seek(0)

                xl_img = OpenpyxlImage(img_stream)
                xl_img.width = target_w
                xl_img.height = target_h

                ws_plates.add_image(xl_img, f"D{r}")
                ws_plates.row_dimensions[r].height = 36
            except Exception as e:
                cell_img.value = "Lỗi ảnh"
                cell_img.font = font_subtitle
                cell_img.alignment = align_center
                ws_plates.row_dimensions[r].height = 20
        else:
            cell_img.value = "Không có ảnh"
            cell_img.font = font_subtitle
            cell_img.alignment = align_center
            ws_plates.row_dimensions[r].height = 20

    if not license_plates:
        empty_pl = ws_plates.cell(row=7, column=1, value="Không tìm thấy biển số xe nào.")
        empty_pl.font = font_subtitle

    # Column dimensions
    ws_plates.column_dimensions["A"].width = 10
    ws_plates.column_dimensions["B"].width = 24
    ws_plates.column_dimensions["C"].width = 24
    ws_plates.column_dimensions["D"].width = 28

    # ==========================================
    # SHEET 4: LỊCH SỬ XE ĐỖ SAI VỊ TRÍ (MISPARKED LOGS)
    # ==========================================
    ws_misparked: Worksheet = wb.create_sheet(title="Lịch Sử Xe Đỗ Sai Vị Trí")
    ws_misparked.views.sheetView[0].showGridLines = True

    # Title Banner
    t_mp = ws_misparked.cell(row=1, column=1, value="⚠️ DANH SÁCH CHI TIẾT CÁC VỤ VI PHẠM ĐỖ SAI VỊ TRÍ")
    ws_misparked.merge_cells("A1:F2")
    t_mp.font = font_title
    t_mp.fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")
    t_mp.alignment = align_center

    cell_mp_a4 = ws_misparked.cell(row=4, column=1, value=f"Bộ lọc: {filter_desc}")
    cell_mp_a4.font = font_subtitle

    misparked_headers = ["STT", "ID Sự Kiện", "Thời Gian Ghi Nhận", "Vị Trí Ô Liên Quan", "Mã Định Danh Xe", "Trạng Thái"]
    for col_idx, text in enumerate(misparked_headers, start=1):
        cell = ws_misparked.cell(row=6, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = PatternFill(start_color="F43F5E", end_color="F43F5E", fill_type="solid")
        cell.alignment = align_center
        cell.border = border_all

    for idx, ev in enumerate(misparked_events, start=1):
        r = 6 + idx
        vals = [idx, ev["id"], ev["timestamp"], f"Ô {ev['slot_id']}", ev["vehicle_id"] or "", "ĐỖ SAI VỊ TRÍ"]

        for col_idx, val in enumerate(vals, start=1):
            cell = ws_misparked.cell(row=r, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = align_center
            cell.border = border_all
            if r % 2 == 0 and col_idx != 6:
                cell.fill = fill_zebra
            if col_idx == 6:
                cell.font = font_badge_misparked
                cell.fill = fill_badge_misparked

        ws_misparked.row_dimensions[r].height = 20

    if not misparked_events:
        empty_mp = ws_misparked.cell(row=7, column=1, value="Không tìm thấy sự kiện đỗ sai vị trí nào.")
        empty_mp.font = font_subtitle

    for col in ws_misparked.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column if col[0].column is not None else 1)
        ws_misparked.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # Save to file
    wb.save(file_path)
    return file_path
